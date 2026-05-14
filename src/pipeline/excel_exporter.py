"""
excel_exporter.py — Exportación a Excel

Genera un archivo .xlsx con:
  - Hoja 1: Matriz Normativa (resultado final)
  - Hoja 2: Resumen estadístico
  - Hoja 3: Log completo de documentos procesados
"""

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import config
from .matrix_builder import MATRIX_COLUMNS

logger = logging.getLogger(__name__)

# ─── Paleta de colores ────────────────────────────────────────────────────────
_COLOR_HEADER_MAIN   = "1F4E79"   # Azul oscuro
_COLOR_HEADER_ALT    = "2E75B6"   # Azul medio
_COLOR_ROW_EVEN      = "D6E4F0"   # Azul muy claro
_COLOR_ROW_ODD       = "FFFFFF"   # Blanco
_COLOR_ALTA          = "C6EFCE"   # Verde claro
_COLOR_MEDIA         = "FFEB9C"   # Amarillo claro
_COLOR_BAJA          = "FFCCCC"   # Rojo muy claro
_COLOR_SUMMARY_HEADER = "375623"  # Verde oscuro
_COLOR_SUMMARY_ROW   = "E2EFDA"   # Verde muy claro


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _header_font(bold: bool = True) -> Font:
    return Font(name="Calibri", bold=bold, color="FFFFFF", size=11)


def _cell_font() -> Font:
    return Font(name="Calibri", size=10)


def _wrap_align() -> Alignment:
    return Alignment(wrap_text=True, vertical="top")


def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── Hoja 1: Matriz normativa ─────────────────────────────────────────────────

def _write_matrix_sheet(ws, matrix_rows: List[Dict]):
    """Escribe la hoja de matriz normativa con formato."""
    ws.title = "Matriz Normativa"
    ws.sheet_view.showGridLines = False

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Encabezados
    for col_idx, col_name in enumerate(MATRIX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = _header_fill(_COLOR_HEADER_MAIN)
        cell.font   = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 40

    # Filas de datos
    for row_idx, row_data in enumerate(matrix_rows, start=2):
        fill_color = _COLOR_ROW_EVEN if row_idx % 2 == 0 else _COLOR_ROW_ODD

        for col_idx, col_name in enumerate(MATRIX_COLUMNS, start=1):
            value = row_data.get(col_name, "No disponible")
            cell  = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.font      = _cell_font()
            cell.alignment = _wrap_align()
            cell.border    = _thin_border()

            # Hiperlink para URL Oficial
            if col_name == "URL Oficial" and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 50

    # Anchos de columna
    col_widths = {
        "Investigador":              18,
        "País":                      10,
        "Título de la Norma":        45,
        "Tipo de norma":             22,
        "Estado":                    15,
        "Organismo Emisor/Universidad": 35,
        "Dominio":                   20,
        "Vínculo con Educación":     20,
        "Dedicación del Texto":      22,
        "Fecha de Publicación":      18,
        "URL Oficial":               45,
        "Observaciones":             55,
        "Ámbito":                    14,
    }
    for col_idx, col_name in enumerate(MATRIX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions


# ─── Hoja 2: Resumen estadístico ─────────────────────────────────────────────

def _write_summary_sheet(ws, matrix_rows: List[Dict], all_documents: List[Dict]):
    """Escribe la hoja de resumen con estadísticas."""
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    def _h(row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = _header_fill(_COLOR_SUMMARY_HEADER)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _thin_border()
        return cell

    def _d(row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=_COLOR_SUMMARY_ROW)
        cell.font = _cell_font()
        cell.alignment = Alignment(vertical="center")
        cell.border = _thin_border()
        return cell

    current_row = 1

    # Título
    ws.merge_cells(f"A{current_row}:D{current_row}")
    title_cell = ws.cell(row=current_row, column=1,
                         value=f"Resumen — Levantamiento Normativo IA en Educación México")
    title_cell.fill = _header_fill("1F4E79")
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 35
    current_row += 2

    # Metadatos
    _h(current_row, 1, "Investigador"); _d(current_row, 2, config.RESEARCHER_NAME); current_row += 1
    _h(current_row, 1, "País");         _d(current_row, 2, config.COUNTRY);         current_row += 1
    _h(current_row, 1, "Fecha");        _d(current_row, 2, datetime.now().strftime("%Y-%m-%d %H:%M")); current_row += 2

    # ── Estadísticas generales ──
    _h(current_row, 1, "ESTADÍSTICAS GENERALES"); current_row += 1
    _h(current_row, 1, "Métrica"); _h(current_row, 2, "Valor"); current_row += 1

    total_docs       = len(all_documents)
    gov_docs         = sum(1 for d in all_documents if d.get("source_type") == "government")
    uni_docs         = sum(1 for d in all_documents if d.get("source_type") == "university")
    open_docs        = sum(1 for d in all_documents if d.get("source_type") == "open")
    total_normativas = len(matrix_rows)
    alta_count       = sum(1 for d in all_documents if d.get("heuristic_label") == "ALTA")
    media_count      = sum(1 for d in all_documents if d.get("heuristic_label") == "MEDIA")
    baja_count       = sum(1 for d in all_documents if d.get("heuristic_label") == "BAJA")

    stats = [
        ("Total documentos procesados",         total_docs),
        ("  — Fuente gobierno",                  gov_docs),
        ("  — Fuente universidades",             uni_docs),
        ("  — Búsqueda abierta",                 open_docs),
        ("Normativas detectadas (IA/heurística)", total_normativas),
        ("Documentos ALTA relevancia",           alta_count),
        ("Documentos MEDIA relevancia",          media_count),
        ("Documentos BAJA relevancia",           baja_count),
    ]
    for label, val in stats:
        _d(current_row, 1, label); _d(current_row, 2, val); current_row += 1

    current_row += 1

    # ── Por tipo de norma ──
    _h(current_row, 1, "NORMATIVAS POR TIPO"); current_row += 1
    _h(current_row, 1, "Tipo"); _h(current_row, 2, "Cantidad"); current_row += 1
    tipo_counter = Counter(r.get("Tipo de norma", "No especificado") for r in matrix_rows)
    for tipo, count in tipo_counter.most_common():
        _d(current_row, 1, tipo); _d(current_row, 2, count); current_row += 1

    current_row += 1

    # ── Por ámbito ──
    _h(current_row, 1, "NORMATIVAS POR ÁMBITO"); current_row += 1
    _h(current_row, 1, "Ámbito"); _h(current_row, 2, "Cantidad"); current_row += 1
    ambito_counter = Counter(r.get("Ámbito", "No especificado") for r in matrix_rows)
    for ambito, count in ambito_counter.most_common():
        _d(current_row, 1, ambito); _d(current_row, 2, count); current_row += 1

    current_row += 1

    # ── Ranking de universidades ──
    _h(current_row, 1, "RANKING UNIVERSIDADES (por producción normativa)"); current_row += 1
    _h(current_row, 1, "Universidad"); _h(current_row, 2, "Normativas"); current_row += 1
    uni_counter = Counter(
        r.get("_university_name", r.get("Organismo Emisor/Universidad", ""))
        for r in matrix_rows
        if r.get("_source_type") == "university"
    )
    for uni, count in uni_counter.most_common(20):
        if uni:
            _d(current_row, 1, uni); _d(current_row, 2, count); current_row += 1

    # Anchos
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ─── Hoja 3: Log completo ────────────────────────────────────────────────────

def _write_log_sheet(ws, log_rows: List[Dict]):
    """Escribe la hoja de auditoría con todos los documentos procesados."""
    ws.title = "Log Documentos"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    if not log_rows:
        return

    headers = list(log_rows[0].keys())

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _header_fill(_COLOR_HEADER_ALT)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 35

    for row_idx, row_data in enumerate(log_rows, start=2):
        label = row_data.get("Label heurístico", "")
        fill_color = {"ALTA": _COLOR_ALTA, "MEDIA": _COLOR_MEDIA}.get(label, _COLOR_ROW_ODD)

        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.font      = _cell_font()
            cell.alignment = _wrap_align()
            cell.border    = _thin_border()

        ws.row_dimensions[row_idx].height = 35

    # Auto-size columns (approximate)
    for col_idx, header in enumerate(headers, start=1):
        max_len = max(len(header), 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.auto_filter.ref = ws.dimensions


# ─── Exportador principal ─────────────────────────────────────────────────────

def export_to_excel(
    matrix_rows: List[Dict],
    all_documents: List[Dict],
    log_rows: List[Dict],
    output_path: Path = config.OUTPUT_EXCEL,
) -> Path:
    """
    Genera el archivo Excel final con las 3 hojas.

    Args:
        matrix_rows:   filas de la matriz normativa (solo normativas confirmadas)
        all_documents: todos los documentos procesados (para resumen)
        log_rows:      log completo de documentos (para auditoría)
        output_path:   ruta de salida del archivo

    Returns:
        Path del archivo generado.
    """
    logger.info(f"[excel] Generando Excel: {output_path}")
    logger.info(f"[excel] {len(matrix_rows)} normativas | {len(all_documents)} documentos totales")

    wb = openpyxl.Workbook()

    # Hoja 1: Matriz
    ws_matrix = wb.active
    _write_matrix_sheet(ws_matrix, matrix_rows)

    # Hoja 2: Resumen
    ws_summary = wb.create_sheet()
    _write_summary_sheet(ws_summary, matrix_rows, all_documents)

    # Hoja 3: Log
    ws_log = wb.create_sheet()
    _write_log_sheet(ws_log, log_rows)

    # Propiedades del libro
    wb.properties.title    = "Normativa IA Educación México"
    wb.properties.subject  = "Levantamiento normativo automatizado"
    wb.properties.creator  = config.RESEARCHER_NAME
    wb.properties.keywords = "inteligencia artificial, normativa, educación, México"

    wb.save(output_path)
    logger.info(f"[excel] Guardado en: {output_path}")
    return output_path
