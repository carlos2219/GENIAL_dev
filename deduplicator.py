"""
deduplicator.py — Eliminación de duplicados

Estrategia dual:
  1. Normalización de URL (elimina duplicados exactos/casi exactos)
  2. Hash de contenido (elimina páginas distintas con el mismo texto)
"""

import hashlib
import re
from urllib.parse import urlparse, urlunparse, parse_qs, urlencode
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


# ─── Normalización de URL ─────────────────────────────────────────────────────

# Parámetros de tracking que no afectan el contenido
_TRACKING_PARAMS = {
    "utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term",
    "fbclid", "gclid", "ref", "referrer", "source", "_ga", "mc_cid",
}


def normalize_url(url: str) -> str:
    """
    Normaliza una URL para comparación:
    - Convierte a minúsculas
    - Elimina esquema http/https diferencias (unifica a https)
    - Quita trailing slashes
    - Elimina parámetros de tracking
    - Quita fragmentos (#...)
    """
    try:
        url = url.strip()
        parsed = urlparse(url.lower())

        # Unificar esquema
        scheme = "https"
        netloc = parsed.netloc.replace("www.", "")

        # Limpiar parámetros de tracking
        params = parse_qs(parsed.query, keep_blank_values=False)
        clean_params = {k: v for k, v in params.items() if k not in _TRACKING_PARAMS}
        clean_query = urlencode(clean_params, doseq=True)

        path = parsed.path.rstrip("/") or "/"
        normalized = urlunparse((scheme, netloc, path, "", clean_query, ""))
        return normalized
    except Exception:
        return url.lower().strip()


def _content_hash(text: str) -> str:
    """Hash SHA-256 de los primeros 2000 caracteres del texto (detecta clones de contenido)."""
    sample = re.sub(r"\s+", " ", text[:2000]).strip().lower()
    return hashlib.sha256(sample.encode("utf-8", errors="replace")).hexdigest()


# ─── Deduplicación ───────────────────────────────────────────────────────────

def deduplicate(documents: List[Dict]) -> List[Dict]:
    """
    Recibe lista de documentos y devuelve la lista sin duplicados.

    Priorización cuando hay colisión:
      - Se prefiere el documento con mayor heuristic_score
      - En empate, se prefiere el que tiene más texto extraído
    """
    # Paso 1: deduplicar por URL normalizada
    by_url: Dict[str, Dict] = {}
    for doc in documents:
        norm = normalize_url(doc.get("url", ""))
        if norm not in by_url:
            by_url[norm] = doc
        else:
            # Mantener el de mayor score
            existing = by_url[norm]
            if doc.get("heuristic_score", 0) > existing.get("heuristic_score", 0):
                by_url[norm] = doc
            elif len(doc.get("extracted_text", "")) > len(existing.get("extracted_text", "")):
                by_url[norm] = doc

    deduped_by_url = list(by_url.values())
    logger.info(f"[dedup] URL: {len(documents)} → {len(deduped_by_url)}")

    # Paso 2: deduplicar por hash de contenido (sólo si hay texto)
    by_content: Dict[str, Dict] = {}
    no_text: List[Dict] = []

    for doc in deduped_by_url:
        text = doc.get("extracted_text", "")
        if not text or len(text) < 100:
            no_text.append(doc)
            continue
        chash = _content_hash(text)
        if chash not in by_content:
            by_content[chash] = doc
        else:
            existing = by_content[chash]
            if doc.get("heuristic_score", 0) > existing.get("heuristic_score", 0):
                by_content[chash] = doc

    result = list(by_content.values()) + no_text
    logger.info(f"[dedup] contenido: {len(deduped_by_url)} → {len(result)}")
    return result


def remove_url_duplicates_only(documents: List[Dict]) -> List[Dict]:
    """
    Versión rápida (solo URL) para usar antes de la extracción de contenido.
    Evita descargar el mismo documento dos veces.
    """
    seen: set = set()
    result: List[Dict] = []
    for doc in documents:
        norm = normalize_url(doc.get("url", ""))
        if norm not in seen:
            seen.add(norm)
            result.append(doc)
    logger.info(f"[dedup-url] {len(documents)} → {len(result)}")
    return result


def load_known_urls_from_excel(excel_path: str) -> set:
    """
    Carga las URLs de la hoja 'Registro de Normativa' del Excel de la matriz
    definitiva y las retorna como un conjunto de URLs normalizadas.

    Úsalas como skip-list en pre-extracción para no re-procesar documentos
    que ya forman parte de la matriz final.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_name = "Registro de Normativa"
        if sheet_name not in wb.sheetnames:
            logger.warning(f"[dedup] Hoja '{sheet_name}' no encontrada en {excel_path}")
            return set()

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return set()

        headers = [str(h).strip() if h else "" for h in rows[0]]
        # Buscar columna de URL flexible (URL Oficial, url_oficial, URL, etc.)
        url_col = None
        for candidate in ("URL Oficial", "url_oficial", "URL", "Url Oficial"):
            if candidate in headers:
                url_col = headers.index(candidate)
                break

        if url_col is None:
            logger.warning("[dedup] No se encontró columna de URL en el Excel de matriz")
            return set()

        known: set = set()
        for row in rows[1:]:
            url = row[url_col] if row[url_col] else ""
            if url and str(url).strip():
                known.add(normalize_url(str(url).strip()))

        logger.info(f"[dedup] {len(known)} URLs conocidas cargadas desde {excel_path}")
        return known

    except ImportError:
        logger.warning("[dedup] openpyxl no instalado; no se puede cargar skip-list del Excel")
        return set()
    except Exception as e:
        logger.warning(f"[dedup] Error cargando Excel skip-list: {e}")
        return set()


def filter_known_urls(documents: List[Dict], known_urls: set) -> List[Dict]:
    """
    Elimina de la lista los documentos cuya URL normalizada
    ya existe en el conjunto known_urls (skip-list del Excel definitivo).
    """
    if not known_urls:
        return documents
    filtered = [d for d in documents if normalize_url(d.get("url", "")) not in known_urls]
    skipped = len(documents) - len(filtered)
    if skipped:
        logger.info(f"[dedup] {skipped} documentos omitidos (ya en matriz definitiva)")
    return filtered
